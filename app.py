import streamlit as st
import openpyxl
import os
from datetime import datetime

file_name = "missions.xlsx"

# Create the Excel file if it doesn't exist
if not os.path.exists(file_name):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Missions"
    ws.append(["Mission", "Date", "Time"])
    ws_finished = wb.create_sheet("Finished")
    ws_finished.append(["Mission", "Date", "Time", "Finished At"])
    wb.save(file_name)

def load_missions():
    wb = openpyxl.load_workbook(file_name)
    ws = wb["Missions"]
    missions = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        missions.append(row)
    return missions

def save_mission(mission, date, time):
    wb = openpyxl.load_workbook(file_name)
    ws = wb["Missions"]
    ws.append([mission, date, time])
    wb.save(file_name)

def delete_mission(mission):
    wb = openpyxl.load_workbook(file_name)
    ws = wb["Missions"]
    for row in ws.iter_rows(min_row=2):
        if row[0].value == mission:
            ws.delete_rows(row[0].row)
            break
    wb.save(file_name)

def load_finished():
    wb = openpyxl.load_workbook(file_name)
    ws = wb["Finished"]
    finished = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        finished.append(row)
    return finished

# Streamlit UI
st.title("📌 Mission Reminder App")

st.header("إضافة مهمة جديدة")
with st.form("add_mission"):
    mission = st.text_input("المهمة")
    date = st.date_input("التاريخ")
    time = st.text_input("الوقت (ساعة:دقيقة)")
    submit = st.form_submit_button("إضافة المهمة")
    if submit and mission and time:
        save_mission(mission, date.strftime("%Y-%m-%d"), time)
        st.success(f"تمت إضافة المهمة '{mission}'")

st.header("قائمة المهام")
missions = load_missions()
if missions:
    for m in missions:
        col1, col2 = st.columns([3,1])
        with col1:
            st.write(f"**{m[0]}** | 📅 {m[1]} | 🕒 {m[2]}")
        with col2:
            if st.button("حذف", key=f"delete_{m[0]}"):
                delete_mission(m[0])
                st.experimental_rerun()
else:
    st.info("لا توجد مهام حالياً")

st.header("المهام المنجزة")
finished = load_finished()
if finished:
    for f in finished:
        st.write(f"✅ {f[0]} | 📅 {f[1]} | 🕒 {f[2]} | انتهت في: {f[3]}")
else:
    st.info("لا توجد مهام منجزة بعد")
